import email     # Import the 'email' library to work with email messages and MIME structures
import imaplib   # Import the 'imaplib' library to interact with the IMAP protocol for email retrieval
from email.header import decode_header  # Import 'decode_header' to decode email headers and 'parsedate_to_datetime' to parse date strings
from email.utils import parsedate_to_datetime
import html2text  # Import 'html2text' to convert HTML content to plain text
import json   # Import 'json' for JSON encoding and decoding
import fitz  #PyMuPDF   # Import 'fitz' from PyMuPDF to handle PDF documents
import io   # Import 'io' for working with input/output streams
from concurrent import futures   # Import 'concurrent.futures' for concurrent execution using ThreadPoolExecutor
import openpyxl   # Import 'openpyxl' to work with Excel files
from docx import Document   # Import 'Document' from 'docx' for handling Word documents
import time  # Import 'time' for measuring execution time


# Function to retrieve unseen emails from Gmail using IMAP
def get_emails(imapUserEmail, imapPassword):
    imap_server = "imap.gmail.com"
    imap_port = 993

    # Connect to the Gmail IMAP server over SSL
    imap_conn = imaplib.IMAP4_SSL(imap_server, imap_port)

    # Log in to the Gmail account
    imap_conn.login(imapUserEmail, imapPassword)

    # Select the 'inbox' folder
    imap_conn.select('inbox')

    # Search for unseen (UNSEEN) emails
    _, message_ids = imap_conn.search(None, "UNSEEN")

    # List to store email data
    emails = []

    # Iterate over unseen email message IDs
    for message_id in message_ids[0].split():
        try:
            # Fetch the raw email data for the given message ID
            _, data = imap_conn.fetch(message_id, "(RFC822)")
            raw_email = data[0][1]

            # Parse the raw email data into an EmailMessage object
            email_message = email.message_from_bytes(raw_email)

            # Extract key information from the email
            subject = email_message["Subject"]
            sender = email.utils.parseaddr(email_message["From"])[0]
            body1 = ""
            attachments = []
            date_received = parsedate_to_datetime(email_message["Date"])

            # Check if the email is multipart (contains different parts like text and attachments)
            if email_message.is_multipart():
                for part in email_message.walk():
                    # Extract plain text content
                    if part.get_content_type() == "text/plain":
                        body1 = part.get_payload(decode=True).decode()
                    # Convert HTML content to plain text
                    elif part.get_content_type() == "text/html":
                        body1 = html2text.html2text(part.get_payload(decode=True).decode())
                    # Extract attachments
                    elif part.get_content_maintype() != 'multipart' and part.get('Content-Disposition') is not None:
                        filename = part.get_filename()
                        if filename:
                            filename, encoding = decode_header(filename)[0]
                            if isinstance(filename, bytes):
                                filename = filename.decode(encoding or "utf-8")
                            attachment_data = part.get_payload(decode=True)
                            attachments.append({
                                "filename": filename,
                                "data": attachment_data
                            })

            # Create a dictionary with email data and append it to the list
            email_data = {
                "subject": subject,
                "sender": sender,
                "body": body1,
                "date_received": date_received,
                "email_id": message_id,
                "attachments": attachments
            }
            emails.append(email_data)
        except Exception as e:
            # Print an error message and mark the email as seen (read)
            print(f"Error processing email: {str(e)}")
            imap_conn.store(message_id, '-FLAGS', '(\Seen)')

    # Close the IMAP connection
    imap_conn.close()

    # Return the list of email data dictionaries
    return emails


# Function to extract text content from a PDF attachment
def extract_text_from_pdf(attachment_data):
    try:
        pdf_document = fitz.open(stream=attachment_data, filetype="pdf")
        text_content = ""
        for page_number in range(pdf_document.page_count):
            page = pdf_document[page_number]
            text_content += page.get_text()
        return text_content
    except Exception as e:
        # Print an error message if extraction fails and return an empty string
        print(f"Error extracting text from PDF: {str(e)}")
        return ""


# Function to extract text content from a Word document attachment
def extract_text_from_word(attachment_data):
    try:
        doc = Document(io.BytesIO(attachment_data))
        text_content = ""
        for paragraph in doc.paragraphs:
            text_content += paragraph.text + "\n"
        return text_content.strip()
    except Exception as e:
        # Print an error message if extraction fails and return an empty string
        print(f"Error extracting text from Word document: {str(e)}")
        return ""


# Function to extract text content from an Excel spreadsheet attachment
def extract_text_from_excel(attachment_data):
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(attachment_data), read_only=True)
        text_content = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    text_content += str(cell.value) + " "
        return text_content
    except Exception as e:
        # Print an error message if extraction fails and return an empty string
        print(f"Error extracting text from Excel spreadsheet: {str(e)}")
        return ""


# Function to process email data and print relevant information
def op_process_mail(email_data):
    print("Subject:", email_data["subject"])
    print("Sender:", email_data["sender"])
    print("Date Received:", email_data["date_received"])

    if email_data["attachments"]:
        for attachment in email_data["attachments"]:
            filename = attachment["filename"]
            content = attachment["data"]

            # Determine the type of attachment and extract text content accordingly
            if filename.lower().endswith('.pdf'):
                pdf_text_content = extract_text_from_pdf(content)
                attachment_info = {
                    "filename": filename,
                    "content_type": "text_from_pdf",
                    "content": pdf_text_content
                }
            elif filename.lower().endswith('.docx'):
                docx_text_content = extract_text_from_word(content)
                attachment_info = {
                    "filename": filename,
                    "content_type": "text_from_docx",
                    "content": docx_text_content
                }
            elif filename.lower().endswith(('.xlsx', '.xls')):
                excel_text_content = extract_text_from_excel(content)
                attachment_info = {
                    "filename": filename,
                    "content_type": "text_from_excel",
                    "content": excel_text_content
                }
            elif filename.lower().endswith('.py'):
                python_script_content = content.decode('utf-8')
                attachment_info = {
                    "filename": filename,
                    "content_type": "text_from_python",
                    "content": python_script_content
                }
            else:
                attachment_info = {
                    "filename": filename,
                    "content_type": "unknown",
                    "content": "Cannot extract text from this file type"
                }

            # Print attachment information or a message indicating no attachments
            if attachment_info:
                print("Attachments:")
                print(json.dumps(attachment_info, indent=2))
            else:
                print("No Attachments")

           # Print email body if available, otherwise, indicate no body content
            if email_data["body"]:
               print("Body:", email_data["body"])



# Function to run the processing of emails concurrently using ThreadPoolExecutor
def run_parallel(emails):
    with futures.ThreadPoolExecutor() as executor:
        executor.map(op_process_mail, emails)

# Record the start time
start = time.time()

# Gmail account credentials
imapUserEmail = "jayshri@vrdella.com"
imapPassword = "ochx uqsr loga iixw"

# Retrieve unseen emails and process them concurrently
emails = get_emails(imapUserEmail, imapPassword)
run_parallel(emails)

# Record the end time
end = time.time()

# Print the total execution time
print("Total Execution Time:", end - start)
